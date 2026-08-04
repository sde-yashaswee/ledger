import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import os
import json
import pandas as pd
from parser import parse_ledger

class LedgerPro:
    def __init__(self, root):
        self.root = root
        self.root.title("Ledger Pro")
        self.root.geometry("720x750")
        self.root.configure(bg="#1E1E1E")  # Dark Background
        
        # Modern Palette
        self.colors = {
            "bg": "#1E1E1E",
            "card": "#2D2D2D",
            "accent": "#007AFF", # Apple Blue
            "text": "#FFFFFF",
            "text_dim": "#AAAAAA",
            "success": "#34C759",
            "warning": "#FF9F0A",
            "error": "#FF453A",
            "border": "#3D3D3D"
        }

        self.input_file = ""
        self.processed_data = None

        self.create_widgets()

    def create_widgets(self):
        # --- Header ---
        header_frame = tk.Frame(self.root, bg=self.colors["bg"])
        header_frame.pack(pady=(40, 20), padx=50, fill='x')
        
        tk.Label(header_frame, text="Ledger Pro", font=("Helvetica", 32, "bold"), 
                 bg=self.colors["bg"], fg=self.colors["text"]).pack(anchor='w')
        tk.Label(header_frame, text="Converts Busy Exports to Programmable Formats", 
                 font=("Helvetica", 14), bg=self.colors["bg"], fg=self.colors["text_dim"]).pack(anchor='w', pady=(5, 0))

        # --- Selection Card ---
        self.card = tk.Frame(self.root, bg=self.colors["card"], bd=0, highlightthickness=1, highlightbackground=self.colors["border"])
        self.card.pack(pady=20, padx=50, fill='x')
        
        self.file_label = tk.Label(self.card, text="Select a .htm ledger file...", 
                                   font=("Helvetica", 12), bg=self.colors["card"], fg=self.colors["text_dim"], 
                                   padx=20, pady=25)
        self.file_label.pack(side=tk.LEFT, fill='x', expand=True)

        self.browse_btn = tk.Button(self.card, text="Browse", command=self.select_file, 
                                    font=("Helvetica", 11, "bold"), bg=self.colors["accent"], fg="white",
                                    highlightthickness=0, bd=0, padx=20, pady=10, cursor="hand2")
        self.browse_btn.pack(side=tk.RIGHT, padx=15)

        # --- Action Button ---
        self.process_btn = tk.Button(self.root, text="Analyze & Verify", command=self.process_file, 
                                     state=tk.DISABLED, font=("Helvetica", 14, "bold"), 
                                     bg=self.colors["accent"], fg="white", disabledforeground="#555555",
                                     highlightthickness=0, bd=0, pady=12, width=25, cursor="hand2")
        self.process_btn.pack(pady=10)

        # --- Summary Panel ---
        summary_label_frame = tk.Frame(self.root, bg=self.colors["bg"])
        summary_label_frame.pack(fill='x', padx=55, pady=(20, 5))
        tk.Label(summary_label_frame, text="AUDIT REPORT", font=("Helvetica", 10, "bold"), 
                 bg=self.colors["bg"], fg=self.colors["accent"]).pack(anchor='w')

        self.summary_box = tk.Frame(self.root, bg=self.colors["card"], highlightthickness=1, highlightbackground=self.colors["border"])
        self.summary_box.pack(padx=50, fill='both', expand=True)

        # Grid-based clean summary
        self.stats_frame = tk.Frame(self.summary_box, bg=self.colors["card"], pady=20, padx=20)
        self.stats_frame.pack(fill='x')
        
        # Row 0: Status + View Errors Button
        tk.Label(self.stats_frame, text="Status:", font=("Helvetica", 11), bg=self.colors["card"], fg=self.colors["text_dim"]).grid(row=0, column=0, sticky='w', pady=5)
        
        status_container = tk.Frame(self.stats_frame, bg=self.colors["card"])
        status_container.grid(row=0, column=1, sticky='e', pady=5)
        
        self.status_lbl = tk.Label(status_container, text="Waiting...", font=("Helvetica", 11, "bold"), bg=self.colors["card"], fg=self.colors["text"])
        self.status_lbl.pack(side=tk.LEFT)
        
        self.view_errors_btn = tk.Button(status_container, text="View Discrepancies", command=self.show_errors, 
                                         font=("Helvetica", 9, "bold"), bg=self.colors["error"], fg="white",
                                         highlightthickness=0, bd=0, padx=8, pady=2, cursor="hand2")
        # Hidden initially
        
        # Other Rows
        self.debit_lbl = self.create_stat_row(self.stats_frame, "Debit Reconciled:", "---", 1)
        self.credit_lbl = self.create_stat_row(self.stats_frame, "Credit Reconciled:", "---", 2)
        self.balance_lbl = self.create_stat_row(self.stats_frame, "Net Balance:", "---", 3)
        self.records_lbl = self.create_stat_row(self.stats_frame, "Total Entries:", "---", 4)
        
        self.stats_frame.columnconfigure(1, weight=1)

        # --- Export Actions ---
        self.actions_frame = tk.Frame(self.root, bg=self.colors["bg"])
        self.actions_frame.pack(pady=30, padx=50, fill='x')
        
        self.json_btn = tk.Button(self.actions_frame, text="Export JSON", command=self.save_json, 
                                  state=tk.DISABLED, font=("Helvetica", 11, "bold"), 
                                  bg=self.colors["card"], fg=self.colors["text"],
                                  highlightthickness=1, highlightbackground=self.colors["border"], 
                                  bd=0, pady=10, width=18)
        self.json_btn.pack(side=tk.LEFT, padx=(0, 10), expand=True, fill='x')

        self.excel_btn = tk.Button(self.actions_frame, text="Export Excel", command=self.save_excel, 
                                   state=tk.DISABLED, font=("Helvetica", 11, "bold"), 
                                   bg=self.colors["card"], fg=self.colors["text"],
                                   highlightthickness=1, highlightbackground=self.colors["border"], 
                                   bd=0, pady=10, width=18)
        self.excel_btn.pack(side=tk.LEFT, padx=(10, 0), expand=True, fill='x')

    def create_stat_row(self, parent, label, value, row):
        tk.Label(parent, text=label, font=("Helvetica", 11), bg=self.colors["card"], fg=self.colors["text_dim"]).grid(row=row, column=0, sticky='w', pady=5)
        val_lbl = tk.Label(parent, text=value, font=("Helvetica", 11, "bold"), bg=self.colors["card"], fg=self.colors["text"])
        val_lbl.grid(row=row, column=1, sticky='e', pady=5)
        return val_lbl

    def select_file(self):
        filename = filedialog.askopenfilename(
            title="Select Ledger HTML File",
            filetypes=(("HTML files", "*.htm *.html"), ("All files", "*.*"))
        )
        if filename:
            self.input_file = filename
            short_name = os.path.basename(filename)
            if len(short_name) > 35: short_name = short_name[:32] + "..."
            self.file_label.config(text=short_name, fg=self.colors["text"])
            self.process_btn.config(state=tk.NORMAL)
            self.reset_ui()

    def reset_ui(self):
        self.status_lbl.config(text="Ready to analyze", fg=self.colors["text"])
        self.view_errors_btn.pack_forget()
        self.debit_lbl.config(text="---", fg=self.colors["text"])
        self.credit_lbl.config(text="---", fg=self.colors["text"])
        self.balance_lbl.config(text="---", fg=self.colors["text"])
        self.records_lbl.config(text="---", fg=self.colors["text"])
        self.json_btn.config(state=tk.DISABLED)
        self.excel_btn.config(state=tk.DISABLED)
        self.processed_data = None

    def process_file(self):
        try:
            self.processed_data = parse_ledger(self.input_file)
            v = self.processed_data['verification']

            # Update Summary Panel
            status_text = "PASSED" if v['success'] else "DISCREPANCY"
            status_color = self.colors["success"] if v['success'] else self.colors["warning"]
            
            self.status_lbl.config(text=status_text, fg=status_color)
            
            if not v['success']:
                self.view_errors_btn.pack(side=tk.LEFT, padx=(10, 0))
            else:
                self.view_errors_btn.pack_forget()
            
            match_deb = "MATCHED" if v['Totals_Reconciled'] else "MISMATCH"
            match_cre = "MATCHED" if v['Totals_Reconciled'] else "MISMATCH"
            
            self.debit_lbl.config(text=f"{match_deb} ({v['Adjusted_Debit_Sum']:,.2f})", 
                                  fg=self.colors["success"] if v['Totals_Reconciled'] else self.colors["error"])
            self.credit_lbl.config(text=f"{match_cre} ({v['Adjusted_Credit_Sum']:,.2f})", 
                                   fg=self.colors["success"] if v['Totals_Reconciled'] else self.colors["error"])
            
            self.balance_lbl.config(text=v['Final_Balance'])
            self.records_lbl.config(text=str(len(self.processed_data['data'])))

            self.json_btn.config(state=tk.NORMAL)
            self.excel_btn.config(state=tk.NORMAL)
            
        except Exception as e:
            messagebox.showerror("Process Error", str(e))

    def show_errors(self):
        if not self.processed_data or not self.processed_data['errors']:
            messagebox.showinfo("No Errors", "No specific row discrepancies found.")
            return
            
        err_win = tk.Toplevel(self.root)
        err_win.title("Discrepancy Report")
        err_win.geometry("800x400")
        err_win.configure(bg=self.colors["bg"])
        
        frame = tk.Frame(err_win, bg=self.colors["bg"], padx=20, pady=20)
        frame.pack(fill='both', expand=True)
        
        tk.Label(frame, text="Row-Level Discrepancies", font=("Helvetica", 14, "bold"), 
                 bg=self.colors["bg"], fg=self.colors["error"]).pack(anchor='w', pady=(0, 15))
        
        # Table
        columns = ("SR", "Date", "Expected", "Actual", "Diff")
        tree = ttk.Treeview(frame, columns=columns, show='headings', height=10)
        
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100, anchor='center')
            
        tree.pack(fill='both', expand=True)
        
        # Scrollbar
        sb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        sb.pack(side='right', fill='y')
        tree.configure(yscrollcommand=sb.set)
        
        for err in self.processed_data['errors']:
            tree.insert("", "end", values=(
                err.get("Serial Number", "N/A"),
                err.get("Date", "N/A"),
                f"{err.get('Expected_Balance', 0):,.2f}",
                f"{err.get('Sheet_Balance', 0):,.2f}",
                f"{err.get('Difference', 0):,.2f}"
            ))

    def save_json(self):
        if not self.processed_data: return
        path = filedialog.asksaveasfilename(defaultextension=".json", initialfile=os.path.basename(self.input_file).replace(".htm", ".json"))
        if path:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(self.processed_data, f, indent=4)
            messagebox.showinfo("Export Successful", "JSON file saved.")

    def save_excel(self):
        if not self.processed_data: return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=os.path.basename(self.input_file).replace(".htm", ".xlsx"))
        if path:
            try:
                from openpyxl.styles import Alignment, Font, PatternFill
                df = pd.DataFrame(self.processed_data['data'])
                
                with pd.ExcelWriter(path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Ledger')
                    worksheet = writer.sheets['Ledger']
                    
                    # Map column names to 1-based indices
                    cols = {cell.value: i+1 for i, cell in enumerate(worksheet[1])}
                    
                    date_col = cols.get("Date")
                    part_col = cols.get("Particulars")
                    deb_col = cols.get("Debit")
                    cre_col = cols.get("Credit")
                    bal_col = cols.get("Balance")
                    
                    # Alignment and Values calculation
                    total_deb = 0
                    total_cre = 0
                    
                    for row_idx in range(2, worksheet.max_row + 1):
                        debit = worksheet.cell(row=row_idx, column=deb_col).value
                        credit = worksheet.cell(row=row_idx, column=cre_col).value
                        
                        try:
                            deb_val = float(debit) if debit else 0
                        except:
                            deb_val = 0
                        try:
                            cre_val = float(credit) if credit else 0
                        except:
                            cre_val = 0
                            
                        total_deb += deb_val
                        total_cre += cre_val
                            
                        if deb_val > 0:
                            align = Alignment(horizontal='left')
                        elif cre_val > 0:
                            align = Alignment(horizontal='right')
                        else:
                            align = Alignment(horizontal='center')
                            
                        if date_col:
                            worksheet.cell(row=row_idx, column=date_col).alignment = align
                        if part_col:
                            worksheet.cell(row=row_idx, column=part_col).alignment = align

                    # Add Totals Row
                    last_data_row = worksheet.max_row
                    total_row = last_data_row + 2
                    
                    worksheet.cell(row=total_row, column=1, value="TOTAL")
                    worksheet.cell(row=total_row, column=deb_col, value=total_deb)
                    worksheet.cell(row=total_row, column=cre_col, value=total_cre)
                    worksheet.cell(row=total_row, column=bal_col, value=round(total_deb - total_cre, 2))
                    
                    # Style Totals Row
                    blue_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    bold_font = Font(bold=True)
                    
                    for c in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=total_row, column=c)
                        cell.fill = blue_fill
                        cell.font = bold_font

                messagebox.showinfo("Export Successful", "Excel file saved with Totals.")
            except Exception as e:
                messagebox.showerror("Excel Error", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    # High DPI support for Mac/Windows
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except:
        pass
    app = LedgerPro(root)
    root.mainloop()
